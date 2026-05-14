"""
Lit le tableur v6 normalisé et génère les JSON consommés par le site statique.

Sortie dans `observations-presidentielles-2027/data/` :
 - candidats.json   (compatible site v2)
 - votes-cles.json  (votes-clés AN/Sénat/PE-emblématiques, format v2)
 - votes-bruts.json (NOUVEAU : drill-down PE par sujet)
 - financement.json (compatible site v2)
 - sources.json     (compatible site v2)
 - sujets.json      (depuis T_sujets, compatible site v2)
 - glossaire.json   (conservé statique — généré par extract_xlsx.py legacy ou
                    pré-existant ; si absent ici, copié depuis legacy)

Usage : py observations-presidentielles-2027/scripts/build_site_data.py
"""
from __future__ import annotations
import json
import sys
import re
from collections import defaultdict
from pathlib import Path

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent  # remonte hors du site
V6 = ROOT / 'analyse_candidats_2027_v6_normalise.xlsx'
SITE_DATA = ROOT / 'observations-presidentielles-2027' / 'data'

CANDIDAT_SLUGS_BY_ID = {1: 'bardella', 2: 'philippe', 3: 'retailleau', 4: 'melenchon', 5: 'attal'}
SUJET_ID_TO_KEY = {  # pour positions_key dans T_sujets
    1: 'immigration', 2: 'climat', 3: 'fiscalite', 4: 'retraites',
    5: 'europe', 6: 'securite', 7: 'ukraine', 8: None,
}


def load_table(wb, name):
    """Renvoie une liste de dicts pour une table v6."""
    ws = wb[name]
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def index_by(rows, key):
    return {r[key]: r for r in rows}


def candidat_label_for(slug):
    return {
        'bardella': 'Jordan Bardella',
        'philippe': 'Édouard Philippe',
        'retailleau': 'Bruno Retailleau',
        'melenchon': 'Jean-Luc Mélenchon',
        'attal': 'Gabriel Attal',
    }[slug]


def parti_label_for(slug):
    return {
        'bardella': 'Rassemblement National',
        'philippe': 'Horizons',
        'retailleau': 'Les Républicains',
        'melenchon': 'La France Insoumise',
        'attal': 'Renaissance',
    }[slug]


def initiales(slug):
    return {
        'bardella': 'JB', 'philippe': 'EP', 'retailleau': 'BR',
        'melenchon': 'JLM', 'attal': 'GA',
    }[slug]


def fmt_eur(v):
    if v is None:
        return None
    try:
        v = int(v)
    except (ValueError, TypeError):
        return None
    if v >= 1_000_000:
        return f'{v/1_000_000:.2f} M€'.replace('.', ',')
    if v >= 1_000:
        return f'{v/1_000:.0f} k€'
    return f'{v} €'


def main():
    print(f'Lecture {V6.name}...')
    wb = openpyxl.load_workbook(V6, read_only=True, data_only=True)

    candidats = load_table(wb, 'T_candidats')
    partis = load_table(wb, 'T_partis')
    sujets = load_table(wb, 'T_sujets')
    sources = load_table(wb, 'T_sources')
    indicateurs = load_table(wb, 'T_indicateurs')
    textes_votes = load_table(wb, 'T_textes_votes')
    mandats = load_table(wb, 'T_mandats')
    votes_bruts = load_table(wb, 'T_votes_bruts')
    votes_agreges = load_table(wb, 'T_votes_agreges')
    valeurs = load_table(wb, 'T_valeurs')
    affaires = load_table(wb, 'T_affaires')
    positions = load_table(wb, 'T_positions')
    fin_camp = load_table(wb, 'T_finances_campagnes')
    fin_part = load_table(wb, 'T_finances_partis')

    print(f'  Lignes : candidats={len(candidats)} sujets={len(sujets)} '
          f'mandats={len(mandats)} valeurs={len(valeurs)} '
          f'votes_bruts={len(votes_bruts)} textes={len(textes_votes)}')

    # Index par id
    sources_by_id = index_by(sources, 'id')
    sujets_by_id = index_by(sujets, 'id')
    indicateurs_by_id = index_by(indicateurs, 'id')
    textes_by_id = index_by(textes_votes, 'id')
    partis_by_id = index_by(partis, 'id')

    def src_url(sid):
        if sid is None:
            return None
        s = sources_by_id.get(sid)
        return s['url'] if s else None

    def src_label(sid):
        if sid is None:
            return None
        s = sources_by_id.get(sid)
        return s['label'] if s else None

    def payload_from_valeur(v):
        """Construit un payload {value, source_url, source_label} pour le site."""
        if v is None:
            return None
        if v['valeur_num'] is not None:
            ind = indicateurs_by_id.get(v['indicateur_id'])
            unite = v['unite']
            if unite == 'EUR':
                value = fmt_eur(v['valeur_num'])
            elif unite == 'pct':
                value = f"{v['valeur_num']*100:.1f} %".replace('.', ',')
            else:
                value = str(v['valeur_num'])
            if ind and ind.get('nom_court') and v.get('periode_debut') and v.get('periode_fin'):
                # ajouter contexte temporel
                pass
        else:
            value = v.get('valeur_text') or ''
        payload = {'value': value}
        if v.get('source_id'):
            payload['source_url'] = src_url(v['source_id'])
            payload['source_label'] = src_label(v['source_id'])
        return payload

    # ----- Construction par candidat -----
    cand_out = {}
    for c in candidats:
        slug = CANDIDAT_SLUGS_BY_ID.get(c['id'])
        if not slug:
            continue
        # Identité
        identite = {
            'date_naissance': {'value': str(c['date_naissance']) if c['date_naissance'] else None, 'source_url': 'https://fr.wikipedia.org/', 'source_label': 'Wikipédia'},
            'lieu_naissance': {'value': c['lieu_naissance'], 'source_url': 'https://fr.wikipedia.org/'},
            'formation': {'value': c['formation_initiale']},
            'profession_origine': {'value': c['profession_origine']},
            'parti': {'value': parti_label_for(slug)},
        }
        # Mandats du candidat
        c_mandats = [m for m in mandats if m['candidat_id'] == c['id']]
        # Pour la rétro-compat, recrée 5 champs typiques de l'onglet identité v5
        institution_to_field = {
            'Mandat actuel': 'mandat_principal',
            'Première élection': 'premiere_election',
            'Fonctions gouvernementales': 'fonctions_gouvernementales',
            'Présidence de parti': 'presidence_parti',
        }
        for m in c_mandats:
            field = institution_to_field.get(m['institution'])
            if field:
                identite[field] = {
                    'value': m['intitule'],
                    'source_url': src_url(m['source_id']),
                    'source_label': src_label(m['source_id']),
                }
        # Candidature 2027 : récupérer depuis T_positions ou via une valeur dédiée
        # On la prend depuis le mandat "Candidature 2027" ou défaut "Pressenti"
        cand_2027_value = None
        cand_2027_src = None
        for m in c_mandats:
            if 'candidature' in (m['intitule'] or '').lower() or '2027' in (m['intitule'] or ''):
                cand_2027_value = m['intitule']
                cand_2027_src = m['source_id']
                break
        if not cand_2027_value:
            # Fallback : statut par défaut depuis notes du candidat
            note = c.get('notes') or ''
            if 'présidentielle 2027' in note.lower() or 'candidat à la présidentielle' in note.lower():
                # Extract une phrase
                m_sentence = re.search(r'([^.]*candidat[^.]*2027[^.]*)', note, re.IGNORECASE)
                cand_2027_value = m_sentence.group(1).strip() if m_sentence else 'Pressenti'
            else:
                cand_2027_value = 'Pressenti / non officiellement déclaré'
        identite['candidature_2027'] = {
            'value': cand_2027_value,
            'source_url': src_url(cand_2027_src),
            'source_label': src_label(cand_2027_src),
        }

        # Activité parlementaire (depuis T_valeurs filtrées)
        activite = {}
        v_candidat = [v for v in valeurs if v['candidat_id'] == c['id']]
        # Mapping indicateur_id -> clé v2
        ind_to_key = {
            1: 'presence_hemicycle', 2: 'presence_hemicycle',
            3: 'presence_commissions', 5: 'cohesion_groupe',
            6: 'interventions', 7: 'interventions',
            8: 'propositions_loi', 10: 'amendements',
            13: 'rapports', 14: 'questions', 15: 'questions',
            33: 'conference_presidents_pe',
        }
        for v in v_candidat:
            key = ind_to_key.get(v['indicateur_id'])
            if not key:
                continue
            p = payload_from_valeur(v)
            if p:
                # Premier vu gagne (on ne concatène pas)
                if key not in activite:
                    activite[key] = p

        # Patrimoine
        patrimoine = {}
        ind_to_key_pat = {
            16: 'patrimoine_declare', 18: 'patrimoine_declare',
            20: 'revenus_annuels',
        }
        for v in v_candidat:
            key = ind_to_key_pat.get(v['indicateur_id'])
            if key and key not in patrimoine:
                p = payload_from_valeur(v)
                if p:
                    patrimoine[key] = p

        # Positions publiques (depuis T_positions)
        c_positions = [p for p in positions if p['candidat_id'] == c['id']]
        positions_out = {}
        for p in c_positions:
            sid = p['sujet_id']
            if sid is None:
                continue
            sujet = sujets_by_id.get(sid)
            if not sujet:
                continue
            pos_key = sujet.get('positions_key') or sujet['slug']
            if pos_key in positions_out:
                continue  # premier vu gagne
            positions_out[pos_key] = {
                'value': p['enonce'],
                'source_url': src_url(p['source_id']),
                'source_label': src_label(p['source_id']),
            }
        # Tribunes / manifestations
        for p in c_positions:
            if p['type'] == 'tribune' and 'tribunes' not in positions_out:
                positions_out['tribunes'] = {
                    'value': p['enonce'],
                    'source_url': src_url(p['source_id']),
                    'source_label': src_label(p['source_id']),
                }
            elif p['type'] == 'manifestation' and 'manifestations' not in positions_out:
                positions_out['manifestations'] = {
                    'value': p['enonce'],
                    'source_url': src_url(p['source_id']),
                    'source_label': src_label(p['source_id']),
                }

        # Affaires
        c_affaires = [a for a in affaires if a['candidat_id'] == c['id']]
        affaires_out = [{
            'intitule': a['intitule'],
            'nature': a['nature'],
            'statut_juridique': a['statut_juridique'],
            'annee': a['notes'],
            'source_url': src_url(a['source_principale_id']),
            'source_label': src_label(a['source_principale_id']),
        } for a in c_affaires]

        # Financement campagne
        c_campagnes = [f for f in fin_camp if f['candidat_id'] == c['id']]
        financement = {}
        if c_campagnes:
            f0 = c_campagnes[0]
            financement['derniere_campagne'] = {
                'value': f0['election'],
                'source_url': src_url(f0['source_id']),
            }
            financement['total_depenses'] = {
                'value': fmt_eur(f0['depenses_eur']) if f0['depenses_eur'] else None,
                'source_url': src_url(f0['source_id']),
            }

        # Synthèse (vide pour l'instant — placeholder)
        synthese = {}

        cand_out[slug] = {
            'slug': slug,
            'nom': c['nom_complet'],
            'parti': parti_label_for(slug),
            'initiales': initiales(slug),
            'identite': identite,
            'activite_parlementaire': activite,
            'patrimoine': patrimoine,
            'positions': positions_out,
            'financement': financement,
            'synthese': synthese,
            'affaires': affaires_out,
            'discours_vs_actes': [],  # non encore migré (optionnel)
        }

    # ----- Bardella PE stats annuelles depuis T_votes_agreges -----
    bardella_pe_annual = []
    bardella_agr = [a for a in votes_agreges if a['candidat_id'] == 1]
    by_year = defaultdict(lambda: {'pour': 0, 'contre': 0, 'abstention': 0, 'absent': 0, 'votes_totaux': 0})
    for a in bardella_agr:
        y = a['annee']
        by_year[y]['pour'] += a['n_pour'] or 0
        by_year[y]['contre'] += a['n_contre'] or 0
        by_year[y]['abstention'] += a['n_abstention'] or 0
        by_year[y]['absent'] += a['n_absent'] or 0
        by_year[y]['votes_totaux'] += a['n_total'] or 0
    for y in sorted(by_year):
        d = by_year[y]
        d['annee'] = y
        d['presence_pct'] = round(100 * (d['votes_totaux'] - d['absent']) / max(d['votes_totaux'], 1), 1) if d['votes_totaux'] else None
        bardella_pe_annual.append(d)
    cand_out['bardella']['bardella_pe_stats'] = {'annual': bardella_pe_annual}

    # Bardella PE votes : les 24 emblématiques de l'onglet v5 (chambre=PE et annee>=2019, filtrer 1ers)
    # Plus simple : tous les votes PE Bardella seront dans votes-bruts.json ; conserver la liste sélective des 24
    sample_bardella = [t for t in textes_votes if t['chambre'] == 'PE' and t.get('ref_officielle') and not str(t['ref_officielle']).startswith('HTV:')][:24]
    cand_out['bardella']['bardella_pe_votes'] = [{
        'theme': '', 'date': t['date_vote'], 'titre': t['titre_long'],
        'position': next((v['position'] for v in votes_bruts if v['candidat_id'] == 1 and v['texte_vote_id'] == t['id']), 'N_APPLICABLE'),
        'position_raw': None, 'groupe': '', 'aligne': '',
    } for t in sample_bardella]

    # ----- candidats.json -----
    candidats_json = {
        'metadata': {
            'date_arrete': '2026-05-12',
            'source_fichier': V6.name,
            'n_candidats': len(cand_out),
        },
        'candidats': cand_out,
    }
    (SITE_DATA / 'candidats.json').write_text(
        json.dumps(candidats_json, ensure_ascii=False, indent=2, default=str),
        encoding='utf-8'
    )
    print(f'  data/candidats.json  OK')

    # ----- votes-cles.json -----
    # Format : tableau de textes votés non-PE-mass (les 20 emblématiques v5 + textes PE référencés v5)
    # Critère : textes dont ref_officielle ne commence pas par HTV: (donc pas issus de HowTheyVote mass)
    votes_cles_textes = [t for t in textes_votes if not str(t.get('ref_officielle') or '').startswith('HTV:')]
    votes_cles = []
    for t in votes_cles_textes:
        positions_for_text = {}
        for slug, slug_id in {v: k for k, v in CANDIDAT_SLUGS_BY_ID.items()}.items():
            vb = next((v for v in votes_bruts if v['candidat_id'] == slug_id and v['texte_vote_id'] == t['id']), None)
            if vb:
                positions_for_text[slug] = {
                    'position': vb['position'].replace('N_APPLICABLE', 'N/A') if vb['position'] else 'N/A',
                    'detail': vb.get('detail_textuel') or None,
                    'source_url': src_url(vb['source_id']),
                    'source_label': src_label(vb['source_id']),
                }
            else:
                positions_for_text[slug] = {'position': 'N/A', 'detail': None}
        sujet = sujets_by_id.get(t.get('sujet_principal_id'))
        votes_cles.append({
            'theme': sujet['titre'] if sujet else '',
            'texte': t['titre_long'] or t['titre_court'],
            'annee': t['annee'],
            'source_label': None,
            'positions': positions_for_text,
        })
    (SITE_DATA / 'votes-cles.json').write_text(
        json.dumps(votes_cles, ensure_ascii=False, indent=2, default=str),
        encoding='utf-8'
    )
    print(f'  data/votes-cles.json OK ({len(votes_cles)} textes)')

    # ----- votes-bruts.json (drill-down) -----
    # Pour chaque candidat, pour chaque sujet, liste les votes bruts (texte + position + date + source)
    bruts_by_candidat = defaultdict(lambda: defaultdict(list))
    for v in votes_bruts:
        slug = CANDIDAT_SLUGS_BY_ID.get(v['candidat_id'])
        if not slug:
            continue
        t = textes_by_id.get(v['texte_vote_id'])
        if not t:
            continue
        sujet = sujets_by_id.get(t.get('sujet_principal_id'))
        sujet_slug = sujet['slug'] if sujet else 'autre'
        bruts_by_candidat[slug][sujet_slug].append({
            'texte': t['titre_long'] or t['titre_court'],
            'annee': t.get('annee'),
            'date': v.get('date_vote') or t.get('date_vote'),
            'chambre': t.get('chambre'),
            'ref': t.get('ref_officielle'),
            'position': v['position'],
        })
    # Tri par date desc dans chaque liste
    for slug, by_sujet in bruts_by_candidat.items():
        for sujet_slug in by_sujet:
            by_sujet[sujet_slug].sort(key=lambda x: str(x.get('date') or ''), reverse=True)
    (SITE_DATA / 'votes-bruts.json').write_text(
        json.dumps(bruts_by_candidat, ensure_ascii=False, indent=2, default=str),
        encoding='utf-8'
    )
    total_bruts = sum(sum(len(v) for v in d.values()) for d in bruts_by_candidat.values())
    print(f'  data/votes-bruts.json OK ({total_bruts} votes bruts au total)')

    # ----- financement.json -----
    financement_json = {
        'campagnes_par_candidat': [{
            'candidat_nom': next((c['nom_complet'] for c in candidats if c['id'] == f['candidat_id']), ''),
            'election': f['election'],
            'date_decision': f['date_decision_cnccfp'],
            'depenses': fmt_eur(f['depenses_eur']),
            'recettes': fmt_eur(f['recettes_eur']),
            'remboursement': fmt_eur(f['remboursement_eur']) or 'Admis au remboursement',
            'reformations': f['reformations_detail'] or fmt_eur(f['reformations_eur']),
            'source_url': src_url(f['source_id']),
        } for f in fin_camp],
        'comptes_partis_2024': [{
            'parti': partis_by_id.get(p['parti_id'], {}).get('nom_long', ''),
            'total_bilan': fmt_eur(p['total_bilan_eur']),
            'produits_2024': fmt_eur(p['produits_eur']),
            'aide_publique_2024': fmt_eur(p['aide_publique_eur']),
            'aide_publique_2025': None,
            'dettes_total': fmt_eur(p['dettes_total_eur']),
            'dont_banques': fmt_eur(p['dettes_banques_eur']),
            'dont_personnes_physiques': fmt_eur(p['dettes_pers_physiques_eur']),
        } for p in fin_part],
        'partis_source_label': 'CNCCFP — Avis publication générale des comptes des partis 2024 (février 2026)',
        'partis_source_url': 'https://cnccfp.fr/wp-content/uploads/2026/02/CNCCFP_DP_Avis_2024.pdf',
    }
    (SITE_DATA / 'financement.json').write_text(
        json.dumps(financement_json, ensure_ascii=False, indent=2, default=str),
        encoding='utf-8'
    )
    print(f'  data/financement.json OK')

    # ----- sources.json -----
    sources_json = [{
        'label': s['label'], 'url': s['url'], 'organisation': s['organisation'],
        'type': s['type'], 'fiabilite': s['fiabilite'],
    } for s in sources]
    (SITE_DATA / 'sources.json').write_text(
        json.dumps(sources_json, ensure_ascii=False, indent=2),
        encoding='utf-8'
    )
    print(f'  data/sources.json    OK ({len(sources_json)} sources)')

    # ----- sujets.json -----
    sujets_json = []
    for s in sujets:
        sujets_json.append({
            'id': s['slug'],
            'titre': s['titre'],
            'type': s['type'],
            'icon_letter': s['titre'][0].lower(),
            'teasing': s.get('teasing', ''),
            'definition': s.get('definition', ''),
            'positions_key': s.get('positions_key', '') or s['slug'],
            'discours_themes': (s.get('discours_themes') or '').split(',') if s.get('discours_themes') else [],
            'votes_filter': {
                'themes': (s.get('votes_filter_themes') or '').split(',') if s.get('votes_filter_themes') else [],
                'textes_includes': [],
            },
        })
    (SITE_DATA / 'sujets.json').write_text(
        json.dumps(sujets_json, ensure_ascii=False, indent=2),
        encoding='utf-8'
    )
    print(f'  data/sujets.json     OK ({len(sujets_json)} sujets)')

    # glossaire.json est conservé tel quel
    print()
    print(f'Build terminé. Source : {V6.name}')


if __name__ == '__main__':
    main()
