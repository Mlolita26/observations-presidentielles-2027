"""
Extract analyse_candidats_2027_v5.xlsx -> 4 JSON files in data/.

Each cell's value and (when present) its Excel comment are captured.
Comments are the source-of-truth provenance for each datum.

Usage : py scripts/extract_xlsx.py
"""
from __future__ import annotations
import json
import os
import re
import sys
from collections import OrderedDict
from pathlib import Path

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT.parent / 'analyse_candidats_2027_v5.xlsx'
DATA = ROOT / 'data'
DATA.mkdir(exist_ok=True)

CANDIDATS = OrderedDict([
    ('bardella', {'nom': 'Jordan Bardella', 'parti': 'Rassemblement National', 'initiales': 'JB'}),
    ('philippe', {'nom': 'Édouard Philippe', 'parti': 'Horizons', 'initiales': 'EP'}),
    ('retailleau', {'nom': 'Bruno Retailleau', 'parti': 'Les Républicains', 'initiales': 'BR'}),
    ('melenchon', {'nom': 'Jean-Luc Mélenchon', 'parti': 'La France Insoumise', 'initiales': 'JLM'}),
    ('attal', {'nom': 'Gabriel Attal', 'parti': 'Renaissance', 'initiales': 'GA'}),
])
# Default column order for sheets where row N has Bardella in B, Philippe in C, ...
DEFAULT_COL_ORDER = ['bardella', 'philippe', 'retailleau', 'melenchon', 'attal']

URL_RE = re.compile(r'https?://[^\s)]+')
FIAB_RE = re.compile(r'(?:Fiabilit[ée]\s*)?(🟢|🟡|🔴|haute|moyenne|faible)', re.IGNORECASE)
POSITION_MAP = [
    (r'^\s*✅\s*POUR\b|\bPOUR\b', 'POUR'),
    (r'^\s*❌\s*CONTRE\b|\bCONTRE\b', 'CONTRE'),
    (r'^\s*⚪\s*ABSTENTION\b|\bABSTENTION\b', 'ABSTENTION'),
    (r'^\s*⛔\s*ABSENT\b|\bABSENT\b', 'ABSENT'),
    (r'^\s*N/A\b|^\s*—|^\s*-', 'N/A'),
]

stats = {'cells_with_value': 0, 'cells_with_source': 0, 'cells_without_source': 0, 'unique_urls': set()}


GENERIC_SOURCE_URL = [
    ('hatvp', 'https://www.hatvp.fr/'),
    ('cnccfp', 'https://www.cnccfp.fr/'),
    ('nosdéputés', 'https://www.nosdeputes.fr/'),
    ('nosdeputes', 'https://www.nosdeputes.fr/'),
    ('nossénateurs', 'https://www.nossenateurs.fr/'),
    ('nossenateurs', 'https://www.nossenateurs.fr/'),
    ('datan', 'https://www.datan.fr/'),
    ('howtheyvote', 'https://howtheyvote.eu/'),
    ('parlement européen', 'https://www.europarl.europa.eu/'),
    ('europarl', 'https://www.europarl.europa.eu/'),
    ('assemblée nationale', 'https://www.assemblee-nationale.fr/'),
    ('sénat', 'https://www.senat.fr/'),
    ('legifrance', 'https://www.legifrance.gouv.fr/'),
    ('légifrance', 'https://www.legifrance.gouv.fr/'),
    ('wikipédia', 'https://fr.wikipedia.org/'),
    ('wikipedia', 'https://fr.wikipedia.org/'),
    ('le point', 'https://www.lepoint.fr/'),
    ('libération', 'https://www.liberation.fr/'),
    ('france 24', 'https://www.france24.com/'),
    ('franceinfo', 'https://www.franceinfo.fr/'),
    ('public sénat', 'https://www.publicsenat.fr/'),
    ('le figaro', 'https://www.lefigaro.fr/'),
    ('politis', 'https://www.politis.fr/'),
    ('dalloz', 'https://www.dalloz-actualite.fr/'),
    ('jorf', 'https://www.legifrance.gouv.fr/'),
]


def extract_url(text):
    if not text:
        return None
    m = URL_RE.search(text)
    return m.group(0).rstrip('.,;)') if m else None


def fallback_generic_url(text):
    """Map common source labels (HATVP, CNCCFP, Datan...) to a generic landing URL."""
    if not text:
        return None
    low = text.lower()
    for needle, url in GENERIC_SOURCE_URL:
        if needle in low:
            return url
    return None


def cell_payload(cell, fallback_url=None, fallback_label=None):
    """Return {value, source_url, source_label, coord} or None if empty.

    fallback_url/label : used when the cell has no own comment but a shared
    row-level source column (typically column G or I) carries the provenance.
    """
    v = cell.value
    cmt = cell.comment.text if cell.comment else None
    if (v is None or (isinstance(v, str) and not v.strip())) and not cmt:
        return None
    own_url = extract_url(cmt) if cmt else None
    own_generic = fallback_generic_url(cmt) if (cmt and not own_url) else None
    out = {
        'value': v if not isinstance(v, str) else v.strip(),
        'coord': cell.coordinate,
    }
    # Also try generic mapping on the cell's own value if it mentions HATVP / CNCCFP etc.
    own_value_generic = fallback_generic_url(v) if (isinstance(v, str) and not own_url and not own_generic) else None
    url = own_url or own_generic or fallback_url or own_value_generic
    if url:
        out['source_url'] = url
        stats['unique_urls'].add(url)
    if cmt:
        label = URL_RE.sub('', cmt).strip(' .;:|-\n\t')
        label = re.sub(r'\s+', ' ', label)
        out['source_label'] = label[:400]
    elif fallback_label:
        out['source_label'] = fallback_label[:400]
    if v is not None:
        stats['cells_with_value'] += 1
        if url:
            stats['cells_with_source'] += 1
        else:
            stats['cells_without_source'] += 1
    return out


def normalize_position(raw_value):
    """Map a cell value from votes-cles sheet to a normalized position code."""
    if raw_value is None:
        return 'N/A', None
    s = str(raw_value).strip()
    if not s:
        return 'N/A', None
    for pat, code in POSITION_MAP:
        if re.search(pat, s):
            # detail is the rest after the icon/keyword
            detail = re.sub(r'^[✅❌⚪⛔]\s*(POUR|CONTRE|ABSTENTION|ABSENT)\b\s*', '', s, count=1).strip()
            detail = re.sub(r'^(POUR|CONTRE|ABSTENTION|ABSENT|N/A)\b\s*', '', detail, count=1).strip()
            return code, (detail or None)
    return 'AUTRE', s


def row_fallback(row, source_col_index):
    """Read the shared source/reliability cell at row[source_col_index]. Returns (url, label).
    If no explicit URL is found, try to map the label to a generic source landing page."""
    if source_col_index >= len(row):
        return None, None
    sc = row[source_col_index]
    cmt = sc.comment.text if sc.comment else None
    text_parts = []
    if sc.value:
        text_parts.append(str(sc.value))
    if cmt:
        text_parts.append(cmt)
    blob = ' | '.join(text_parts)
    url = extract_url(blob) or fallback_generic_url(blob)
    label = sc.value if isinstance(sc.value, str) else None
    return url, label


def extract_identite(ws):
    """Sheet '1. Carte d'identité' - rows 2..11 are critères, col A=label, B-F=candidats, G=source/fiabilité."""
    out = {slug: OrderedDict() for slug in CANDIDATS}
    field_keys = {
        'Date de naissance': 'date_naissance',
        'Lieu de naissance': 'lieu_naissance',
        'Formation initiale': 'formation',
        "Profession d'origine": 'profession_origine',
        'Parti politique actuel': 'parti',
        'Mandat principal actuel': 'mandat_principal',
        'Première élection': 'premiere_election',
        'Fonctions gouvernementales': 'fonctions_gouvernementales',
        'Présidence de parti': 'presidence_parti',
        'Candidature présidentielle 2027': 'candidature_2027',
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        label = (row[0].value or '').strip() if row[0].value else ''
        key = field_keys.get(label)
        if not key:
            continue
        fallback_url, fallback_label = row_fallback(row, 6)  # col G
        for i, slug in enumerate(DEFAULT_COL_ORDER, start=1):
            payload = cell_payload(row[i], fallback_url=fallback_url, fallback_label=fallback_label)
            if payload is not None:
                out[slug][key] = payload
    return out


def extract_simple_table(ws, header_row, label_to_key, source_col=6):
    """Generic extraction: row header_row maps columns B-F to candidates; col A is label.
    source_col (0-indexed): column carrying the shared source/fiabilité info (default G=6).
    Returns {slug: {key: payload}}."""
    out = {slug: OrderedDict() for slug in CANDIDATS}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        label_cell = row[0]
        label = (label_cell.value or '').strip() if label_cell.value else ''
        if not label:
            continue
        key = label_to_key.get(label) or label_to_key.get(label.rstrip('*').strip())
        if not key:
            # fall back to slugified label
            key = slugify(label)
        fallback_url, fallback_label = row_fallback(row, source_col)
        for i, slug in enumerate(DEFAULT_COL_ORDER, start=1):
            if i >= len(row):
                continue
            payload = cell_payload(row[i], fallback_url=fallback_url, fallback_label=fallback_label)
            if payload is not None:
                out[slug][key] = payload
    return out


def slugify(s):
    s = s.lower()
    s = re.sub(r'[àâä]', 'a', s)
    s = re.sub(r'[éèêë]', 'e', s)
    s = re.sub(r'[îï]', 'i', s)
    s = re.sub(r'[ôö]', 'o', s)
    s = re.sub(r'[ûüù]', 'u', s)
    s = re.sub(r'[ç]', 'c', s)
    s = re.sub(r'[^a-z0-9]+', '_', s)
    return s.strip('_')


def extract_activite(ws):
    label_to_key = {
        'Nature du mandat à analyser': 'nature_mandat',
        'Taux de présence (hémicycle)': 'presence_hemicycle',
        'Taux de présence (commissions)': 'presence_commissions',
        'Nombre d\'interventions en séance': 'interventions',
        'Propositions de loi déposées': 'propositions_loi',
        'Amendements déposés': 'amendements',
        'Rapports rédigés': 'rapports',
        'Questions écrites / orales': 'questions',
        'Cohérence vote avec groupe (%)': 'cohesion_groupe',
        'Activité notable (commentaire qualitatif)': 'activite_notable',
        'Présence Conférence des présidents (PE - président de groupe)': 'conference_presidents_pe',
    }
    return extract_simple_table(ws, header_row=4, label_to_key=label_to_key)


def extract_patrimoine(ws):
    label_to_key = {
        'Déclaration de patrimoine consultable': 'declaration_consultable',
        'Lien direct vers déclarations': 'lien_declarations',
        'Patrimoine déclaré (€) - dernière déclaration': 'patrimoine_declare',
        'Revenus annuels déclarés (€)': 'revenus_annuels',
        'Activités annexes rémunérées': 'activites_annexes',
        'Activités passées notables (conflits potentiels)': 'activites_passees',
        'Participations / actions / SCI': 'participations',
        'Conjoint·e — activité professionnelle (déclarée)': 'conjoint',
        'Évolution patrimoine entrée/sortie de mandat': 'evolution_patrimoine',
    }
    return extract_simple_table(ws, header_row=4, label_to_key=label_to_key)


def extract_positions(ws):
    label_to_key = {
        'Immigration': 'immigration',
        'Europe': 'europe',
        'Climat / écologie': 'climat',
        'Fiscalité': 'fiscalite',
        'Retraites': 'retraites',
        'Sécurité / justice': 'securite',
        'Ukraine / Russie': 'ukraine',
        'Tribunes notables signées (à compléter)': 'tribunes',
        'Manifestations soutenues publiquement': 'manifestations',
    }
    return extract_simple_table(ws, header_row=3, label_to_key=label_to_key)


def extract_financement(ws):
    label_to_key = {
        'Dernière campagne analysable': 'derniere_campagne',
        'Total dépenses dernière campagne': 'total_depenses',
        'Origine des fonds (dons / prêts / partis)': 'origine_fonds',
        'Prêts bancaires étrangers (le cas échéant)': 'prets_etrangers',
        'Compte validé par CNCCFP': 'compte_valide',
        'Financement du parti (dons annuels)': 'financement_parti',
        '→ Comptes CNCCFP à télécharger': 'comptes_telecharger',
    }
    return extract_simple_table(ws, header_row=4, label_to_key=label_to_key)


def extract_synthese(ws):
    label_to_key = {
        'Données parlementaires disponibles': 'donnees_disponibles',
        'Taux de présence (vs moyenne)': 'presence_synth',
        'Cohérence vote / groupe': 'coherence_groupe',
        'Patrimoine (cohérence revenus/déclaration)': 'patrimoine_synth',
        'Affaires judiciaires en cours': 'affaires_synth',
        'Cohérence discours / votes': 'coherence_discours',
        'Transparence financement': 'transparence_financement',
    }
    return extract_simple_table(ws, header_row=4, label_to_key=label_to_key)


def extract_affaires(ws):
    """Sheet '5. Affaires judiciaires' uses row-per-affair, col A = candidat name."""
    name_to_slug = {
        'Jordan Bardella': 'bardella',
        'Édouard Philippe': 'philippe',
        'Bruno Retailleau': 'retailleau',
        'Jean-Luc Mélenchon': 'melenchon',
        'Gabriel Attal': 'attal',
    }
    out = {slug: [] for slug in CANDIDATS}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        nom_cell = row[0]
        nom = (nom_cell.value or '').strip() if nom_cell.value else ''
        slug = name_to_slug.get(nom)
        if not slug:
            continue
        # The "Source" column F (index 5) often carries the URL for the whole row.
        fallback_url, fallback_label = row_fallback(row, 5)
        intitule = cell_payload(row[1], fallback_url=fallback_url, fallback_label=fallback_label)
        nature = cell_payload(row[2], fallback_url=fallback_url, fallback_label=fallback_label)
        statut = cell_payload(row[3], fallback_url=fallback_url, fallback_label=fallback_label)
        annee = cell_payload(row[4], fallback_url=fallback_url, fallback_label=fallback_label)
        source = cell_payload(row[5])
        affaire = {
            'intitule': intitule['value'] if intitule else None,
            'nature': nature['value'] if nature else None,
            'statut_juridique': statut['value'] if statut else None,
            'annee': annee['value'] if annee else None,
            'source_label': (source or {}).get('value'),
        }
        for c in (statut, source, intitule):
            if c and c.get('source_url'):
                affaire['source_url'] = c['source_url']
                break
        out[slug].append(affaire)
    return out


def extract_discours_vs_actes(ws):
    """Sheet '8. Discours vs actes' : A=Candidat, B=Theme, C=Discours, D=Actes."""
    name_to_slug = {
        'Jordan Bardella': 'bardella',
        'Édouard Philippe': 'philippe',
        'Bruno Retailleau': 'retailleau',
        'Jean-Luc Mélenchon': 'melenchon',
        'Gabriel Attal': 'attal',
    }
    out = {slug: [] for slug in CANDIDATS}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        nom = (row[0].value or '').strip() if row[0].value else ''
        slug = name_to_slug.get(nom)
        if not slug:
            continue
        theme = cell_payload(row[1])
        discours = cell_payload(row[2])
        actes = cell_payload(row[3])
        entry = {
            'theme': theme['value'] if theme else None,
            'discours': discours['value'] if discours else None,
            'actes': actes['value'] if actes else None,
        }
        for c in (actes, discours, theme):
            if c and c.get('source_url'):
                entry['source_url'] = c['source_url']
                break
        out[slug].append(entry)
    return out


def extract_votes_cles(ws):
    """Sheet '3. Votes-clés 2020-2026' - rows 5+. A=Theme, B=Texte, C=Année, D-H=positions, I=Source."""
    out = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        if row[1].value is None and row[0].value is None:
            continue
        theme = row[0].value
        texte = row[1].value
        annee = row[2].value
        if not texte:
            continue
        fallback_url, fallback_label = row_fallback(row, 8)  # col I
        positions = OrderedDict()
        for i, slug in enumerate(DEFAULT_COL_ORDER):
            cell = row[3 + i]
            payload = cell_payload(cell, fallback_url=fallback_url, fallback_label=fallback_label)
            if payload is None:
                positions[slug] = {'position': 'N/A', 'detail': None}
                continue
            code, detail = normalize_position(payload.get('value'))
            entry = {'position': code, 'detail': detail or payload.get('value')}
            if payload.get('source_url'):
                entry['source_url'] = payload['source_url']
            if payload.get('source_label'):
                entry['source_label'] = payload['source_label']
            positions[slug] = entry
        source_payload = cell_payload(row[8]) if len(row) > 8 else None
        out.append({
            'theme': theme,
            'texte': texte,
            'annee': annee,
            'source_label': (source_payload or {}).get('value'),
            'positions': positions,
        })
    return out


def extract_bardella_pe_votes(ws):
    """Sheet 'Bardella PE - 24 votes-clés' - rows 5+. A=Thème B=Date C=Ref D=Titre E=Position F=Groupe G=Aligné."""
    out = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        if row[3].value is None:
            continue
        out.append({
            'theme': row[0].value,
            'date': row[1].value.isoformat() if hasattr(row[1].value, 'isoformat') else row[1].value,
            'ref': row[2].value,
            'titre': row[3].value,
            'position': normalize_position(row[4].value)[0] if row[4].value else 'N/A',
            'position_raw': row[4].value,
            'groupe': row[5].value,
            'aligne': row[6].value,
        })
    return out


def extract_bardella_stats(ws):
    """Sheet 'Stats Bardella PE par année' - rows 5-12 are annual presence."""
    annual = []
    for row in ws.iter_rows(min_row=5, max_row=12):
        if row[0].value is None:
            continue
        try:
            annee = int(row[0].value)
        except (ValueError, TypeError):
            continue
        annual.append({
            'annee': annee,
            'votes_totaux': row[1].value,
            'pour': row[2].value,
            'contre': row[3].value,
            'abstention': row[4].value,
            'absent': row[5].value,
            'presence_pct': row[6].value,
        })
    cohesion = {
        'id_2019_2024': {'votes': 17432, 'alignes': 13790, 'pct': 79.1},
        'pfe_2024_present': {'votes': 4226, 'alignes': 3913, 'pct': 92.6},
        'global': {'votes': 21658, 'alignes': 17703, 'pct': 81.7},
    }
    return {'annual': annual, 'cohesion': cohesion}


def extract_complements_v5(ws):
    """Sheet '10. Compléments v5' - mixed sub-tables. Extract by row ranges identified by headers."""
    # Comptes de campagne : rows 6..7 (data); header row 5
    campagnes = []
    for row in ws.iter_rows(min_row=6, max_row=7):
        if row[0].value is None:
            continue
        campagnes.append({
            'candidat_nom': row[0].value,
            'election': row[1].value,
            'date_decision': row[2].value,
            'depenses': row[3].value,
            'recettes': row[4].value,
            'remboursement': row[5].value,
            'reformations': row[6].value,
            'source_url': row[7].value,
        })
    # Partis : rows 12..22 (data); header row 11
    partis = []
    for row in ws.iter_rows(min_row=12, max_row=22):
        if row[0].value is None:
            continue
        partis.append({
            'parti': row[0].value,
            'total_bilan': row[1].value,
            'produits_2024': row[2].value,
            'aide_publique_2024': row[3].value,
            'aide_publique_2025': row[4].value,
            'dettes_total': row[5].value,
            'dont_banques': row[6].value,
            'dont_personnes_physiques': row[7].value,
        })
    partis_source = 'CNCCFP, Avis publication générale des comptes des partis politiques - Exercice 2024 (février 2026)'
    partis_source_url = 'https://cnccfp.fr/wp-content/uploads/2026/02/CNCCFP_DP_Avis_2024.pdf'
    # Activité comparée : rows 30..38 (data); header row 29
    activite_comparee = []
    for row in ws.iter_rows(min_row=30, max_row=38):
        if row[0].value is None:
            continue
        activite_comparee.append({
            'indicateur': row[0].value,
            'philippe': row[1].value,
            'melenchon': row[2].value,
            'retailleau': row[3].value,
            'attal': row[4].value,
            'bardella': row[5].value,
        })
    # Prises de position majeures : rows 45..57
    prises = []
    for row in ws.iter_rows(min_row=45, max_row=57):
        if row[0].value is None:
            continue
        prises.append({
            'candidat': row[0].value,
            'date': row[1].value,
            'format': row[2].value,
            'sujet': row[3].value,
            'source_url': row[4].value,
        })
    # Affaires : rows 62..68
    affaires_complement = []
    for row in ws.iter_rows(min_row=62, max_row=68):
        if row[0].value is None:
            continue
        affaires_complement.append({
            'candidat': row[0].value,
            'affaire': row[1].value,
            'statut': row[2].value,
            'date_decision': row[3].value,
            'source_url': row[4].value,
        })
    return {
        'campagnes': campagnes,
        'partis': partis,
        'partis_source_label': partis_source,
        'partis_source_url': partis_source_url,
        'activite_comparee': activite_comparee,
        'prises_position_majeures': prises,
        'affaires_complement': affaires_complement,
    }


# ----------------------------------------------------------------------------
# Données éditoriales : sujets et glossaire
# ----------------------------------------------------------------------------

def build_sujets():
    """11 sujets : 8 politiques + 3 transverses. Définitions et mappings écrits ici."""
    return [
        {
            "id": "immigration",
            "titre": "Immigration",
            "type": "politique",
            "icon_letter": "i",
            "teasing": "Politique migratoire, frontières, asile, regroupement familial.",
            "definition": "L'immigration recouvre les politiques d'accueil, de visas, d'asile, de regroupement familial et de gestion des frontières. Les principaux textes 2020-2026 portant sur le sujet sont la loi Asile et Immigration de janvier 2024 (dite « Darmanin ») adoptée à l'Assemblée nationale, et le Pacte UE Asile et Migration adopté au Parlement européen le 10 avril 2024.",
            "positions_key": "immigration",
            "discours_themes": ["Immigration"],
            "votes_filter": {"themes": ["Immigration"], "textes_includes": ["immigration", "asile", "pacte"]}
        },
        {
            "id": "climat",
            "titre": "Climat & écologie",
            "type": "politique",
            "icon_letter": "c",
            "teasing": "Climat, énergie, nucléaire, pesticides, restauration nature.",
            "definition": "Le sujet « climat & écologie » englobe la politique énergétique (nucléaire, renouvelables), la lutte contre le changement climatique et la protection de la biodiversité. Textes-clés : loi Climat et Résilience 2021, loi d'accélération des énergies renouvelables 2023, loi relance du nucléaire 2023, Loi européenne sur le climat (UE).",
            "positions_key": "climat",
            "discours_themes": ["Climat"],
            "votes_filter": {"themes": ["Climat", "Énergie"], "textes_includes": ["climat", "nucléaire", "renouvelable", "énergie"]}
        },
        {
            "id": "economie",
            "titre": "Économie & fiscalité",
            "type": "politique",
            "icon_letter": "€",
            "teasing": "Impôts, dépense publique, plein emploi, industrie.",
            "definition": "Politiques fiscales, budgétaires et industrielles. Les textes-clés 2020-2026 incluent les lois de finances annuelles (PLF 2024, PLF 2025), la loi Plein emploi (2023) et la loi Industrie verte (2023).",
            "positions_key": "fiscalite",
            "discours_themes": ["Pouvoir d'achat"],
            "votes_filter": {"themes": ["Fiscalité", "Travail", "Économie"], "textes_includes": ["PLF", "emploi", "industrie"]}
        },
        {
            "id": "retraites",
            "titre": "Retraites",
            "type": "politique",
            "icon_letter": "r",
            "teasing": "Âge de départ, durée de cotisation, financement.",
            "definition": "La réforme des retraites du printemps 2023 (recul de l'âge de départ à 64 ans) a été adoptée par engagement de responsabilité du gouvernement (49.3) à l'Assemblée nationale le 16 mars 2023, puis validée au Sénat. Aucun nouveau texte structurant depuis.",
            "positions_key": "retraites",
            "discours_themes": [],
            "votes_filter": {"themes": ["Retraites"], "textes_includes": ["retraite"]}
        },
        {
            "id": "europe",
            "titre": "Europe & souveraineté",
            "type": "politique",
            "icon_letter": "e",
            "teasing": "Construction européenne, souveraineté, État de droit, AI Act.",
            "definition": "Position face à la construction européenne, son cadre institutionnel, l'État de droit dans les États membres (Article 7 Hongrie), et les grands textes UE 2020-2026 : AI Act, CBAM (taxe carbone aux frontières), Pacte asile, Climate Law, Ukraine Facility.",
            "positions_key": "europe",
            "discours_themes": ["Europe"],
            "votes_filter": {"themes": ["Europe"], "textes_includes": ["UE", "européen", "AI Act", "asile"]}
        },
        {
            "id": "securite",
            "titre": "Sécurité & justice",
            "type": "politique",
            "icon_letter": "s",
            "teasing": "Police, peines, libertés publiques, laïcité.",
            "definition": "Politiques de sécurité publique, de justice pénale et de libertés publiques. Textes 2020-2026 : loi Sécurité globale 2021, loi confortant les principes de la République 2021 (« séparatisme »), Pass sanitaire 2021, Loi d'orientation et de programmation justice 2023.",
            "positions_key": "securite",
            "discours_themes": ["Libertés publiques"],
            "votes_filter": {"themes": ["Libertés", "Justice"], "textes_includes": ["sécurité", "justice", "principes", "pass"]}
        },
        {
            "id": "ukraine",
            "titre": "Ukraine & politique étrangère",
            "type": "politique",
            "icon_letter": "u",
            "teasing": "Soutien à l'Ukraine, sanctions Russie, OTAN.",
            "definition": "Soutien financier et militaire à l'Ukraine, sanctions contre la Russie, position face à l'OTAN. Texte-clé européen : Ukraine Facility, votée au Parlement européen en février 2024 puis amendée en février 2026.",
            "positions_key": "ukraine",
            "discours_themes": [],
            "votes_filter": {"themes": ["Europe"], "textes_includes": ["Ukraine", "Russie"]}
        },
        {
            "id": "societe",
            "titre": "Société",
            "type": "politique",
            "icon_letter": "o",
            "teasing": "IVG dans la Constitution, fin de vie, droits humains.",
            "definition": "Grandes questions sociétales votées 2020-2026 : inscription de l'interruption volontaire de grossesse (IVG) dans la Constitution (Congrès de Versailles, 4 mars 2024) et projet de loi sur la fin de vie (Assemblée nationale, 27 mai 2025).",
            "positions_key": None,
            "discours_themes": [],
            "votes_filter": {"themes": ["Société"], "textes_includes": ["IVG", "fin de vie", "Constitution"]}
        },
        # ---- Sujets transverses ----
        {
            "id": "patrimoine",
            "titre": "Patrimoine & intérêts",
            "type": "transverse",
            "icon_letter": "p",
            "teasing": "Déclarations HATVP, revenus, biens, activités annexes.",
            "definition": "Les responsables politiques sont tenus de déclarer leur patrimoine et leurs intérêts à la Haute Autorité pour la transparence de la vie publique (HATVP) à l'entrée et à la sortie de chaque mandat exécutif. Cette rubrique compare le patrimoine net déclaré, les revenus annuels, les participations et activités annexes des cinq candidats.",
            "positions_key": None,
            "discours_themes": [],
            "votes_filter": None
        },
        {
            "id": "affaires",
            "titre": "Affaires judiciaires",
            "type": "transverse",
            "icon_letter": "j",
            "teasing": "Procédures connues, statut juridique précis.",
            "definition": "Cette rubrique recense les procédures judiciaires connues publiquement pour chaque candidat, avec le statut juridique exact (témoin assisté, mis en examen, condamné en 1ère instance, etc.). La présomption d'innocence est strictement respectée : aucun terme définitif n'est employé tant qu'une condamnation n'est pas devenue définitive.",
            "positions_key": None,
            "discours_themes": [],
            "votes_filter": None
        },
        {
            "id": "financement",
            "titre": "Financement politique",
            "type": "transverse",
            "icon_letter": "f",
            "teasing": "Comptes de campagne, comptes des partis, CNCCFP.",
            "definition": "Le financement de la vie politique est contrôlé par la Commission nationale des comptes de campagne et des financements politiques (CNCCFP). Cette rubrique présente les comptes de campagne récents analysables et les comptes annuels 2024 des onze principaux partis français.",
            "positions_key": None,
            "discours_themes": [],
            "votes_filter": None
        },
    ]


def build_glossaire():
    """Liste des termes techniques avec définitions courtes et longues."""
    return [
        # ---- Catégorie : Termes juridiques ----
        {"id": "presomption-innocence", "terme": "Présomption d'innocence", "categorie": "Termes juridiques",
         "definition_courte": "Toute personne est présumée innocente tant qu'elle n'a pas été déclarée coupable par une décision de justice définitive.",
         "definition_longue": "Principe fondamental garanti par l'article 9 de la Déclaration des droits de l'homme et l'article 6 §2 de la Convention européenne des droits de l'homme. Tant qu'une condamnation n'est pas devenue définitive (épuisement des voies de recours), il est interdit de présenter une personne mise en cause comme coupable. Ce site n'emploie jamais de termes définitifs (« coupable », « fautif ») avant condamnation définitive."},
        {"id": "enquete-preliminaire", "terme": "Enquête préliminaire", "categorie": "Termes juridiques",
         "definition_courte": "Première phase d'investigation menée par le parquet, avant l'ouverture éventuelle d'une instruction.",
         "definition_longue": "Conduite par le procureur de la République ou la police judiciaire sous son autorité. Elle peut déboucher sur un classement sans suite, une procédure simplifiée, ou l'ouverture d'une information judiciaire (instruction)."},
        {"id": "instruction", "terme": "Instruction judiciaire", "categorie": "Termes juridiques",
         "definition_courte": "Phase d'investigation approfondie menée par un juge d'instruction.",
         "definition_longue": "Le juge d'instruction enquête à charge et à décharge. Il peut entendre les personnes mises en cause sous le statut de simple témoin, de témoin assisté, ou les mettre en examen, avant de décider d'un éventuel renvoi devant la juridiction de jugement."},
        {"id": "temoin-assiste", "terme": "Témoin assisté", "categorie": "Termes juridiques",
         "definition_courte": "Statut intermédiaire entre simple témoin et personne mise en examen. La personne est entendue avec un avocat mais les indices ne suffisent pas à une mise en examen.",
         "definition_longue": "Statut prévu par l'article 113-1 du Code de procédure pénale. La personne est entendue dans le cadre d'une instruction, peut être assistée d'un avocat, mais n'est pas formellement mise en examen — les indices sont jugés insuffisamment graves et concordants. Ce statut est plus favorable que celui de mis en examen ; il n'implique aucune présomption de culpabilité."},
        {"id": "mis-en-examen", "terme": "Mis en examen", "categorie": "Termes juridiques",
         "definition_courte": "Personne contre laquelle existent des indices graves ou concordants rendant vraisemblable sa participation aux faits.",
         "definition_longue": "Article 80-1 du Code de procédure pénale. La mise en examen permet à la personne d'avoir accès à la procédure et de pouvoir faire valoir ses droits de défense. Elle ne constitue en aucun cas une déclaration de culpabilité : la présomption d'innocence demeure pleine et entière."},
        {"id": "renvoye", "terme": "Renvoyé devant la juridiction", "categorie": "Termes juridiques",
         "definition_courte": "Décision du juge d'instruction d'envoyer la personne devant un tribunal pour y être jugée.",
         "definition_longue": "Au terme de l'instruction, le juge peut prononcer un non-lieu (clôture sans suite) ou un renvoi (ordonnance de renvoi devant un tribunal correctionnel ou de mise en accusation devant une cour d'assises). C'est le jugement qui statuera sur la culpabilité."},
        {"id": "condamne-1ere-instance", "terme": "Condamné en 1ère instance", "categorie": "Termes juridiques",
         "definition_courte": "Première décision de condamnation, susceptible d'appel pendant 10 jours (matière correctionnelle).",
         "definition_longue": "La décision n'est pas définitive : la personne ou le ministère public peuvent faire appel, ce qui rejuge l'affaire complètement. La présomption d'innocence partielle subsiste jusqu'à l'exhaussement des voies de recours."},
        {"id": "condamne-appel", "terme": "Condamné en appel", "categorie": "Termes juridiques",
         "definition_courte": "Décision rendue par la cour d'appel après nouvel examen complet du dossier. Susceptible de pourvoi en cassation.",
         "definition_longue": "La cour d'appel rejuge l'affaire en fait et en droit. Sa décision peut être contestée par pourvoi en cassation, mais ce pourvoi ne porte que sur la régularité juridique, pas sur les faits."},
        {"id": "condamne-definitif", "terme": "Condamné définitivement", "categorie": "Termes juridiques",
         "definition_courte": "Condamnation qui ne peut plus être contestée. Toutes les voies de recours sont épuisées.",
         "definition_longue": "La condamnation est devenue irrévocable : pas d'appel possible, ou pourvoi en cassation rejeté/non formé. C'est seulement à ce stade que la personne est juridiquement « coupable »."},
        {"id": "cjr", "terme": "CJR — Cour de justice de la République", "categorie": "Termes juridiques",
         "definition_courte": "Juridiction spéciale compétente pour juger les ministres pour des faits commis dans l'exercice de leurs fonctions.",
         "definition_longue": "Créée en 1993, composée de 12 parlementaires et 3 magistrats. Elle est saisie via une commission des requêtes. Une réforme constitutionnelle visant à sa suppression est régulièrement débattue."},
        {"id": "classement-sans-suite", "terme": "Classement sans suite", "categorie": "Termes juridiques",
         "definition_courte": "Décision du parquet de ne pas poursuivre une plainte ou un signalement.",
         "definition_longue": "Article 40-1 du Code de procédure pénale. Le procureur estime qu'il n'y a pas d'infraction caractérisée, que l'auteur est inconnu, ou que les poursuites ne sont pas opportunes. La victime peut alors se constituer partie civile pour relancer la procédure."},

        # ---- Catégorie : Transparence et déclarations (HATVP) ----
        {"id": "hatvp", "terme": "HATVP", "categorie": "Transparence et déclarations",
         "definition_courte": "Haute Autorité pour la transparence de la vie publique : autorité administrative indépendante chargée du contrôle des déclarations de patrimoine et d'intérêts.",
         "definition_longue": "Créée en 2013 après l'affaire Cahuzac. Elle reçoit, contrôle et rend publiques les déclarations de patrimoine et d'intérêts d'environ 18 000 responsables publics : membres du gouvernement, parlementaires, élus locaux importants, dirigeants d'entreprises publiques, etc."},
        {"id": "dsp", "terme": "DSP — Déclaration de situation patrimoniale", "categorie": "Transparence et déclarations",
         "definition_courte": "Document détaillé des biens (immobilier, comptes, véhicules, dettes) déposé à l'entrée et à la sortie de chaque mandat.",
         "definition_longue": "Distincte de la déclaration d'intérêts. Elle permet de mesurer l'évolution du patrimoine pendant un mandat (et de détecter d'éventuels enrichissements suspects). Publiée sur le site de la HATVP."},
        {"id": "dia", "terme": "DIA — Déclaration d'intérêts et d'activités", "categorie": "Transparence et déclarations",
         "definition_courte": "Document listant les activités professionnelles, mandats, fonctions et intérêts financiers exercés par le déclarant.",
         "definition_longue": "Sert à prévenir et détecter les conflits d'intérêts. Inclut les activités du conjoint (sans valorisation patrimoniale). Mise à jour en cours de mandat si un nouvel intérêt apparaît."},
        {"id": "pantouflage", "terme": "Pantouflage", "categorie": "Transparence et déclarations",
         "definition_courte": "Passage d'un haut fonctionnaire ou d'un responsable politique vers le secteur privé.",
         "definition_longue": "Soumis à un avis de la HATVP (commission de déontologie pour les fonctionnaires). Peut soulever des questions de conflit d'intérêts si le poste privé est lié à un secteur précédemment supervisé."},
        {"id": "conflit-interets", "terme": "Conflit d'intérêts", "categorie": "Transparence et déclarations",
         "definition_courte": "Situation où un intérêt privé peut influencer l'exercice d'une fonction publique.",
         "definition_longue": "Article 2 de la loi du 11 octobre 2013 : « toute situation d'interférence entre un intérêt public et des intérêts publics ou privés qui est de nature à influencer ou paraître influencer l'exercice indépendant, impartial et objectif d'une fonction »."},

        # ---- Catégorie : Financement politique (CNCCFP) ----
        {"id": "cnccfp", "terme": "CNCCFP", "categorie": "Financement politique",
         "definition_courte": "Commission nationale des comptes de campagne et des financements politiques : autorité indépendante contrôlant l'argent en politique.",
         "definition_longue": "Créée en 1990. Elle contrôle les comptes de campagne des candidats et les comptes annuels des partis politiques. Ses décisions peuvent être contestées devant le Conseil d'État. Toutes ses décisions sont publiées au JORF."},
        {"id": "compte-campagne", "terme": "Compte de campagne", "categorie": "Financement politique",
         "definition_courte": "Relevé exhaustif des recettes et dépenses engagées par un candidat pour sa campagne électorale.",
         "definition_longue": "Obligatoire pour toute élection au-dessus d'un certain seuil. Tenu par un mandataire financier (personne physique ou association de financement). Doit être déposé à la CNCCFP dans les 2 mois suivant le scrutin."},
        {"id": "plafond-depenses", "terme": "Plafond de dépenses", "categorie": "Financement politique",
         "definition_courte": "Montant maximum de dépenses qu'un candidat peut engager pendant sa campagne.",
         "definition_longue": "Présidentielle 2022 : 16 851 000 € au 1er tour, 22 509 000 € au 2nd tour. Le dépassement entraîne le rejet du compte et l'inéligibilité du candidat."},
        {"id": "reformation", "terme": "Réformation", "categorie": "Financement politique",
         "definition_courte": "Montant retiré du compte de campagne par la CNCCFP parce qu'il ne respecte pas les règles (justification insuffisante, dépense étrangère à la campagne, etc.).",
         "definition_longue": "La CNCCFP peut réformer (réduire) le compte présenté. Le montant réformé n'est pas remboursé par l'État. Exemples : déplacements personnels, dépenses non liées à la campagne, justificatifs manquants."},
        {"id": "remboursement-public", "terme": "Remboursement public", "categorie": "Financement politique",
         "definition_courte": "Fraction des dépenses remboursée par l'État aux candidats ayant obtenu au moins 5 % des suffrages exprimés.",
         "definition_longue": "Présidentielle : jusqu'à 47,5 % du plafond pour les candidats ayant obtenu ≥5 % au 1er tour. Plus faible pour les autres seuils. Le remboursement est plafonné au montant de l'apport personnel du candidat."},
        {"id": "aide-publique-partis", "terme": "Aide publique aux partis", "categorie": "Financement politique",
         "definition_courte": "Dotation annuelle versée par l'État aux partis politiques (deux fractions : sur résultats législatifs et sur représentation parlementaire).",
         "definition_longue": "Régie par la loi du 11 mars 1988. 1ère fraction : proportionnelle aux suffrages obtenus aux dernières législatives (>1 % dans au moins 50 circonscriptions). 2ème fraction : proportionnelle au nombre de parlementaires affiliés. Total annuel : ~66 M€."},

        # ---- Catégorie : Activité parlementaire ----
        {"id": "scrutin-solennel", "terme": "Scrutin solennel", "categorie": "Activité parlementaire",
         "definition_courte": "Vote public sur l'ensemble d'un texte, en séance plénière, à une date annoncée à l'avance.",
         "definition_longue": "Contrairement au vote à main levée ou au vote ordinaire, le scrutin solennel est nominatif (chaque vote est public) et programmé dans le calendrier. C'est la métrique la plus utilisée pour mesurer la présence des parlementaires."},
        {"id": "scrutin-public", "terme": "Scrutin public", "categorie": "Activité parlementaire",
         "definition_courte": "Vote dont le résultat individuel de chaque parlementaire est rendu public.",
         "definition_longue": "Au Sénat, tous les scrutins publics sont par défaut publics. À l'AN, le scrutin public est demandé par le président d'un groupe ou le président de séance. Les scrutins solennels sont automatiquement publics."},
        {"id": "cohesion-groupe", "terme": "Cohésion de groupe", "categorie": "Activité parlementaire",
         "definition_courte": "Pourcentage de votes d'un parlementaire alignés sur la position majoritaire de son groupe politique.",
         "definition_longue": "Indicateur calculé par Datan.fr (AN) et HowTheyVote.eu (PE). Un score >95 % est typique ; un score <85 % signale un parlementaire « dissident » par rapport à sa famille politique."},
        {"id": "rapporteur", "terme": "Rapporteur (titulaire / fictif)", "categorie": "Activité parlementaire",
         "definition_courte": "Parlementaire chargé d'examiner un texte au fond et de présenter un rapport au reste de l'assemblée.",
         "definition_longue": "« Rapporteur titulaire » : porte officiellement le rapport. « Shadow rapporteur » (rapporteur fictif) au Parlement européen : rapporteur désigné par les autres groupes politiques pour suivre le texte en parallèle. Le nombre de rapports écrits est un indicateur d'implication."},
        {"id": "proposition-loi", "terme": "Proposition de loi (PPL)", "categorie": "Activité parlementaire",
         "definition_courte": "Texte de loi déposé par un parlementaire (à la différence du projet de loi, déposé par le gouvernement).",
         "definition_longue": "Très peu de PPL sont effectivement adoptées (~10 % en moyenne). Le nombre de PPL déposées en tant qu'auteur principal est l'un des indicateurs d'activité législative — mais pas son efficacité."},
        {"id": "projet-loi", "terme": "Projet de loi", "categorie": "Activité parlementaire",
         "definition_courte": "Texte de loi présenté par le gouvernement et examiné par le Parlement.",
         "definition_longue": "Par opposition à la proposition de loi (PPL) déposée par un parlementaire. Les projets de loi représentent la majorité des textes adoptés."},
        {"id": "amendement", "terme": "Amendement", "categorie": "Activité parlementaire",
         "definition_courte": "Proposition de modification d'un texte de loi en discussion.",
         "definition_longue": "Peut être déposé par un parlementaire, un groupe, le rapporteur, ou le gouvernement. Le « taux d'adoption » des amendements (rapport entre amendements adoptés et déposés) est un indicateur d'influence."},
        {"id": "49-3", "terme": "49.3", "categorie": "Activité parlementaire",
         "definition_courte": "Article 49 alinéa 3 de la Constitution : le gouvernement engage sa responsabilité sur un texte, qui est adopté sans vote sauf motion de censure.",
         "definition_longue": "Le Premier ministre engage la responsabilité du gouvernement. Si une motion de censure n'est pas adoptée dans les 24 h par la majorité absolue des députés, le texte est considéré comme adopté. Utilisé notamment pour la réforme des retraites 2023."},
        {"id": "motion-censure", "terme": "Motion de censure", "categorie": "Activité parlementaire",
         "definition_courte": "Texte voté par les députés pour renverser le gouvernement.",
         "definition_longue": "Doit recueillir la majorité absolue des députés (289 voix). Si adoptée, le gouvernement doit démissionner. La motion contre le gouvernement Barnier a été adoptée le 4 décembre 2024 — la première fois depuis 1962."},
        {"id": "conference-presidents", "terme": "Conférence des présidents", "categorie": "Activité parlementaire",
         "definition_courte": "Instance regroupant les présidents de groupes politiques pour organiser le calendrier des travaux d'une assemblée.",
         "definition_longue": "Existe au PE, à l'AN et au Sénat. Y siègent les présidents des groupes politiques. Au Parlement européen, elle fixe l'ordre du jour des sessions plénières."},

        # ---- Catégorie : Parlement européen ----
        {"id": "mep", "terme": "MEP — Member of the European Parliament", "categorie": "Parlement européen",
         "definition_courte": "Député européen, élu pour 5 ans au suffrage universel direct.",
         "definition_longue": "720 députés européens (depuis 2024), répartis par État membre proportionnellement à la population. La France élit 81 députés."},
        {"id": "groupe-id", "terme": "ID — Identity and Democracy", "categorie": "Parlement européen",
         "definition_courte": "Groupe politique du Parlement européen (2019-2024) où siégeait le Rassemblement National et Jordan Bardella.",
         "definition_longue": "Créé en juin 2019, dissous en juillet 2024. Réunissait notamment le RN (France), la Lega (Italie), AfD (Allemagne, ensuite exclue), FPÖ (Autriche). Successeur : groupe Patriots for Europe (PfE)."},
        {"id": "groupe-pfe", "terme": "PfE — Patriots for Europe", "categorie": "Parlement européen",
         "definition_courte": "Groupe politique créé en juillet 2024, présidé par Jordan Bardella. Successeur d'ID.",
         "definition_longue": "Créé le 8 juillet 2024 par Viktor Orbán (Fidesz) et rejoint notamment par le RN, la Lega, FPÖ. Troisième groupe du Parlement européen en nombre de sièges (84 députés)."},
        {"id": "frontex", "terme": "Frontex", "categorie": "Parlement européen",
         "definition_courte": "Agence européenne de garde-frontières et de garde-côtes.",
         "definition_longue": "Basée à Varsovie. Soutient les États membres dans la surveillance des frontières extérieures de l'UE. La « décharge » est le vote par lequel le Parlement européen approuve (ou non) la gestion budgétaire de l'agence."},
        {"id": "ai-act", "terme": "AI Act", "categorie": "Parlement européen",
         "definition_courte": "Règlement européen sur l'intelligence artificielle, adopté en mars 2024.",
         "definition_longue": "Premier cadre juridique mondial sur l'IA. Catégorise les systèmes d'IA selon leur niveau de risque et impose des obligations (transparence, supervision humaine, évaluation). Entrée en vigueur progressive 2025-2027."},
        {"id": "cbam", "terme": "CBAM — Carbon Border Adjustment Mechanism", "categorie": "Parlement européen",
         "definition_courte": "Taxe carbone aux frontières de l'UE sur les importations à fort contenu carbone.",
         "definition_longue": "Vise à éviter les « fuites carbone » (délocalisation vers des pays sans contrainte climatique). S'applique progressivement à partir de 2026 sur l'acier, le ciment, l'aluminium, les engrais, l'électricité et l'hydrogène."},
        {"id": "pacte-asile-ue", "terme": "Pacte UE Asile et Migration", "categorie": "Parlement européen",
         "definition_courte": "Ensemble de 10 textes réformant la politique européenne d'asile et de gestion des frontières, adoptés le 10 avril 2024.",
         "definition_longue": "Prévoit un mécanisme de solidarité obligatoire entre États membres (accueil ou contribution financière), des procédures aux frontières accélérées, et une base de données commune (Eurodac renforcée). Entrée en application : juin 2026."},
        {"id": "ukraine-facility", "terme": "Ukraine Facility", "categorie": "Parlement européen",
         "definition_courte": "Programme européen de soutien financier à l'Ukraine, voté en février 2024 (50 Md€ sur 2024-2027).",
         "definition_longue": "Combine prêts et dons à l'Ukraine pour financer son fonctionnement et sa reconstruction. Voté à la majorité au Parlement européen ; opposition notable du groupe ID puis PfE (dont le RN)."},
    ]


def main():
    print(f'Lecture : {XLSX}')
    if not XLSX.exists():
        print(f'ERREUR : fichier introuvable {XLSX}', file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    identite = extract_identite(wb['1. Carte d\'identité'])
    activite = extract_activite(wb['2. Activité parlementaire'])
    patrimoine = extract_patrimoine(wb['4. Patrimoine et intérêts'])
    positions = extract_positions(wb['6. Positions publiques'])
    financement_sheet = extract_financement(wb['7. Financement'])
    synthese = extract_synthese(wb['9. Synthèse'])
    affaires = extract_affaires(wb['5. Affaires judiciaires'])
    discours = extract_discours_vs_actes(wb['8. Discours vs actes'])
    votes_an = extract_votes_cles(wb['3. Votes-clés 2020-2026'])
    bardella_pe_votes = extract_bardella_pe_votes(wb['Bardella PE - 24 votes-clés'])
    bardella_stats = extract_bardella_stats(wb['Stats Bardella PE par année'])
    complements = extract_complements_v5(wb['10. Compléments v5'])

    # ----- candidats.json -----
    candidats_out = OrderedDict()
    for slug, meta in CANDIDATS.items():
        candidats_out[slug] = {
            'slug': slug,
            'nom': meta['nom'],
            'parti': meta['parti'],
            'initiales': meta['initiales'],
            'identite': identite[slug],
            'activite_parlementaire': activite[slug],
            'patrimoine': patrimoine[slug],
            'positions': positions[slug],
            'financement': financement_sheet[slug],
            'synthese': synthese[slug],
            'affaires': affaires[slug],
            'discours_vs_actes': discours[slug],
        }
    # Inject Bardella PE detail
    candidats_out['bardella']['bardella_pe_votes'] = bardella_pe_votes
    candidats_out['bardella']['bardella_pe_stats'] = bardella_stats

    candidats_json = {
        'metadata': {
            'date_arrete': '2026-05-12',
            'source_fichier': 'analyse_candidats_2027_v5.xlsx',
            'n_candidats': len(CANDIDATS),
        },
        'candidats': candidats_out,
    }
    (DATA / 'candidats.json').write_text(
        json.dumps(candidats_json, ensure_ascii=False, indent=2, default=str),
        encoding='utf-8',
    )
    print(f'   data/candidats.json    OK ({len(candidats_out)} candidats)')

    # ----- votes-cles.json -----
    (DATA / 'votes-cles.json').write_text(
        json.dumps(votes_an, ensure_ascii=False, indent=2, default=str),
        encoding='utf-8',
    )
    print(f'   data/votes-cles.json   OK ({len(votes_an)} textes)')

    # ----- financement.json -----
    financement_out = {
        'campagnes_par_candidat': complements['campagnes'],
        'comptes_partis_2024': complements['partis'],
        'partis_source_label': complements['partis_source_label'],
        'partis_source_url': complements['partis_source_url'],
        'activite_comparee_synthese': complements['activite_comparee'],
        'prises_position_majeures': complements['prises_position_majeures'],
        'affaires_complement': complements['affaires_complement'],
    }
    (DATA / 'financement.json').write_text(
        json.dumps(financement_out, ensure_ascii=False, indent=2, default=str),
        encoding='utf-8',
    )
    print(f'   data/financement.json  OK ({len(complements["partis"])} partis, {len(complements["campagnes"])} campagnes)')

    # ----- sources.json -----
    sources = []
    seen = set()
    # Static sources from "Sources principales" + "Sources et fichiers" sheets
    static = [
        ('HATVP', 'https://www.hatvp.fr', 'Haute Autorité pour la transparence de la vie publique'),
        ('CNCCFP', 'https://www.cnccfp.fr', 'Commission nationale des comptes de campagne et des financements politiques'),
        ('Assemblée nationale', 'https://www.assemblee-nationale.fr', 'AN — scrutins publics, propositions de loi'),
        ('Sénat', 'https://www.senat.fr', 'Sénat — scrutins publics'),
        ('Parlement européen', 'https://www.europarl.europa.eu', 'PE — profils MEPs, activités plénière'),
        ('NosDéputés.fr', 'https://www.nosdeputes.fr', 'Activité parlementaire AN'),
        ('NosSénateurs.fr', 'https://www.nossenateurs.fr', 'Activité parlementaire Sénat'),
        ('Datan.fr', 'https://www.datan.fr', 'Analyses votes AN (cohésion, présence)'),
        ('HowTheyVote.eu', 'https://howtheyvote.eu', 'Votes nominaux Parlement européen'),
    ]
    for label, url, desc in static:
        sources.append({'label': label, 'url': url, 'description': desc, 'type': 'primaire'})
        seen.add(url)
    for url in sorted(stats['unique_urls']):
        if url in seen:
            continue
        sources.append({'label': url.split('/')[2] if '://' in url else url, 'url': url, 'type': 'extraite'})
        seen.add(url)
    (DATA / 'sources.json').write_text(
        json.dumps(sources, ensure_ascii=False, indent=2),
        encoding='utf-8',
    )
    print(f'   data/sources.json      OK ({len(sources)} sources, dont {len(static)} primaires)')

    # ----- sujets.json (éditorial : définitions des 11 sujets) -----
    sujets = build_sujets()
    (DATA / 'sujets.json').write_text(
        json.dumps(sujets, ensure_ascii=False, indent=2),
        encoding='utf-8',
    )
    print(f'   data/sujets.json       OK ({len(sujets)} sujets)')

    # ----- glossaire.json (éditorial : ~40 termes techniques) -----
    glossaire = build_glossaire()
    (DATA / 'glossaire.json').write_text(
        json.dumps(glossaire, ensure_ascii=False, indent=2),
        encoding='utf-8',
    )
    print(f'   data/glossaire.json    OK ({len(glossaire)} termes)')

    # ----- récapitulatif -----
    print()
    print('RÉCAPITULATIF :')
    print(f'  Candidats         : {len(CANDIDATS)}')
    print(f'  Cellules valorisées : {stats["cells_with_value"]}')
    print(f'  Avec source URL   : {stats["cells_with_source"]}')
    print(f'  Sans source URL   : {stats["cells_without_source"]}')
    print(f'  URLs uniques      : {len(stats["unique_urls"])}')
    if stats['cells_with_value']:
        pct = 100 * stats['cells_with_source'] / stats['cells_with_value']
        print(f'  Taux sourçage     : {pct:.1f} %')


if __name__ == '__main__':
    main()
